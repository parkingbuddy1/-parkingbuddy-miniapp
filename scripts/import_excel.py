"""
从Excel重新生成停车场月卡+次卡JSON数据
月卡: 深圳停车场月卡统计图.xlsx
次卡: 深圳停车场次卡套餐明细_甘特图1.xlsx
"""
import json
import openpyxl
import re
import os

BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "cloudfunctions", "parking")

def clean_val(val, default=None):
    if val is None:
        return default
    if isinstance(val, str):
        val = val.strip()
    return val or default

def to_float(val):
    if val is None: return 0
    try: return float(val)
    except: return 0

def to_int(val):
    if val is None: return 0
    try: return int(float(val))
    except: return 0

def sanitize_id(name):
    """生成简短的ID片段"""
    return re.sub(r'[^\w\u4e00-\u9fff]', '', name or '')[:6]

# ============================================================
# PART 1: 月卡数据
# ============================================================
print("=" * 60)
print("处理月卡数据...")
wb = openpyxl.load_workbook(
    r'C:/Users/Administrator/Desktop/深圳停车场月卡统计图（仅月卡比价数据）.xlsx',
    data_only=True
)
ws = wb.active

# Group rows by project name
projects = {}  # name -> {info, packages}
seen_pkg_ids = set()

for row in range(2, ws.max_row + 1):
    name = clean_val(ws.cell(row, 1).value)
    if not name or name == '项目名称':
        continue
    
    city = clean_val(ws.cell(row, 2).value, '深圳市')
    district = clean_val(ws.cell(row, 3).value, '')
    street = clean_val(ws.cell(row, 4).value, '')
    address = clean_val(ws.cell(row, 5).value, '')
    fee_desc = clean_val(ws.cell(row, 6).value, '')
    period = clean_val(ws.cell(row, 7).value, '月卡')
    pkg_type = clean_val(ws.cell(row, 8).value, '全天')
    sub_type = clean_val(ws.cell(row, 9).value, pkg_type)
    pkg_name = clean_val(ws.cell(row, 10).value, '')
    remark = clean_val(ws.cell(row, 11).value, '')
    time_range = clean_val(ws.cell(row, 12).value, '')
    orig_price = to_float(ws.cell(row, 13).value)
    cur_price = to_float(ws.cell(row, 14).value)
    group3 = to_float(ws.cell(row, 15).value)
    group15 = to_float(ws.cell(row, 16).value)
    
    if name not in projects:
        projects[name] = {
            'name': name, 'city': city, 'district': district,
            'street': street.replace('傎', '镇') if street else '',
            'address': address, 'feeDesc': fee_desc,
            'latitude': None, 'longitude': None,
            'packages': []
        }
    else:
        # Update with more complete info if available
        if not projects[name]['feeDesc'] and fee_desc:
            projects[name]['feeDesc'] = fee_desc
        if not projects[name]['address'] and address:
            projects[name]['address'] = address
    
    # Generate unique package ID
    base_id = sanitize_id(name)
    pkg_idx = len(projects[name]['packages'])
    pkg_id = f"pkg_{pkg_idx:02d}_{base_id}"
    
    # Ensure unique
    while pkg_id in seen_pkg_ids:
        pkg_idx += 1
        pkg_id = f"pkg_{pkg_idx:02d}_{base_id}"
    seen_pkg_ids.add(pkg_id)
    
    pkg = {
        '_id': pkg_id,
        'period': period,
        'packageType': pkg_type,
        'subType': sub_type,
        'name': pkg_name or f"{pkg_type}套餐{period}",
        'remark': remark,
        'timeRange': time_range,
        'originalPrice': orig_price,
        'price': cur_price,
        'groupPrice3': group3 if group3 > 0 else cur_price,
        'groupPrice15': group15 if group15 > 0 else cur_price,
        'tags': [period, pkg_type, sub_type]
    }
    projects[name]['packages'].append(pkg)

# Convert to list and assign IDs
monthly_list = []
for idx, (name, proj) in enumerate(sorted(projects.items())):
    proj['_id'] = f"pk_monthly_{idx+1:04d}"
    proj['sort'] = idx + 1
    monthly_list.append(proj)

print(f"月卡项目: {len(monthly_list)} 个, 套餐: {sum(len(p['packages']) for p in monthly_list)} 个")

# Save
path = os.path.join(BASE_DIR, 'parking_lots_monthly.json')
with open(path, 'w', encoding='utf-8') as f:
    json.dump(monthly_list, f, ensure_ascii=False, indent=2)
print(f"已保存: {path}")

# ============================================================
# PART 2: 次卡数据
# ============================================================
print("\n" + "=" * 60)
print("处理次卡数据...")
wb2 = openpyxl.load_workbook(
    r'C:/Users/Administrator/Desktop/深圳停车场次卡套餐明细_甘特图1.xlsx',
    data_only=True
)
ws2 = wb2.active

projects2 = {}
seen_pkg_ids2 = set()

for row in range(2, ws2.max_row + 1):
    name = clean_val(ws2.cell(row, 1).value)
    if not name or name == '项目名称':
        continue
    
    city = clean_val(ws2.cell(row, 2).value, '深圳市')
    district = clean_val(ws2.cell(row, 3).value, '')
    street = clean_val(ws2.cell(row, 4).value, '')
    address = clean_val(ws2.cell(row, 5).value, '')
    pkg_type = clean_val(ws2.cell(row, 6).value, '全天')
    sub_type = clean_val(ws2.cell(row, 7).value, pkg_type)
    pkg_name = clean_val(ws2.cell(row, 8).value, '')
    time_range = clean_val(ws2.cell(row, 9).value, '')
    valid_days = to_int(ws2.cell(row, 10).value)
    in_out = to_int(ws2.cell(row, 11).value)
    orig_price = to_float(ws2.cell(row, 12).value)
    cur_price = to_float(ws2.cell(row, 13).value)
    per_use = to_float(ws2.cell(row, 14).value)
    limit_use = clean_val(ws2.cell(row, 15).value, '不限')
    
    if name not in projects2:
        projects2[name] = {
            'name': name, 'city': city, 'district': district,
            'street': street or '',
            'address': address or '',
            'latitude': None, 'longitude': None,
            'packages': []
        }
    else:
        if not projects2[name]['address'] and address:
            projects2[name]['address'] = address
    
    base_id = sanitize_id(name)
    pkg_idx = len(projects2[name]['packages'])
    pkg_id = f"pkg_{pkg_idx:02d}_{base_id}"
    while pkg_id in seen_pkg_ids2:
        pkg_idx += 1
        pkg_id = f"pkg_{pkg_idx:02d}_{base_id}"
    seen_pkg_ids2.add(pkg_id)
    
    pkg = {
        '_id': pkg_id,
        'period': '次卡',
        'packageType': pkg_type,
        'subType': sub_type,
        'name': pkg_name or f"{pkg_type}套餐",
        'timeRange': time_range,
        'validDays': valid_days,
        'inOutCount': in_out,
        'originalPrice': orig_price,
        'price': cur_price,
        'perUsePrice': per_use if per_use > 0 else cur_price,
        'perUseTimeLimit': time_range,
        'limitType': limit_use if limit_use else '不限',
        'tags': [pkg_type, sub_type, time_range]
    }
    projects2[name]['packages'].append(pkg)

count_list = []
for idx, (name, proj) in enumerate(sorted(projects2.items())):
    proj['_id'] = f"pk_count_{idx+1:04d}"
    proj['sort'] = idx + 1
    count_list.append(proj)

print(f"次卡项目: {len(count_list)} 个, 套餐: {sum(len(p['packages']) for p in count_list)} 个")

path2 = os.path.join(BASE_DIR, 'parking_lots_count.json')
with open(path2, 'w', encoding='utf-8') as f:
    json.dump(count_list, f, ensure_ascii=False, indent=2)
print(f"已保存: {path2}")

print(f"\n{'='*60}")
print(f"全部完成!")
print(f"月卡: {len(monthly_list)} 个项目")
print(f"次卡: {len(count_list)} 个项目")
